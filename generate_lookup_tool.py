#!/usr/bin/env python3
"""
generate_lookup_tool.py

Regenerates index.html for the SME Compensation Matrix Rate Lookup Tool
from the current version of the SME Excel grid.

Run this any time the underlying SME Excel file changes, then re-upload
the freshly generated index.html to your host (Netlify/Vercel/GitHub
Pages/your own server) to replace the old one.

USAGE
-----
    python3 generate_lookup_tool.py "SME_Upload_Grid_UPDATED.xlsx"

Optional: specify a different output filename
    python3 generate_lookup_tool.py "SME_Upload_Grid_UPDATED.xlsx" -o index.html

REQUIREMENTS
------------
    pip install openpyxl

WHAT IT DOES
------------
1. Reads the Excel file's "SME Compensation Matrix" sheet.
2. Expects these columns (in this order — matches the standard template
   plus the Occupancy Code / Occupancy Name columns added for this tool):
       A: Grid ID
       B: Company
       C: Category
       D: Product
       E: Occupancy Code
       F: Occupancy Name
       G: Recivable Net
       H: Payable Net
       (I onward: Start Date, End Date, Created By, etc. — not used by the tool)
3. Builds a nested lookup structure: Company -> Product -> Occupancy Name -> [rate rows]
4. Injects that structure as JSON into index_template.html (which must be
   in the same folder as this script) to produce a new, fully self-contained
   index.html.

NOTE: index_template.html must stay in the same folder as this script —
it is the page shell (design, search logic) with a placeholder where the
data gets inserted. Do not edit the placeholder text __DATA_JSON__ inside it.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "Missing dependency 'openpyxl'.\n"
        "Install it first with:  pip install openpyxl"
    )

SHEET_NAME = "SME Compensation Matrix"
TEMPLATE_FILENAME = "index_template.html"
PLACEHOLDER = "__DATA_JSON__"


def build_lookup_json(xlsx_path: Path) -> str:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(
            f"Sheet '{SHEET_NAME}' not found in {xlsx_path.name}. "
            f"Sheets present: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    data = {}
    companies_set = set()
    row_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 8:
            continue
        _, company, category, product, occ_code, occ_name, recv, pay = row[:8]

        if not company or not product:
            continue

        companies_set.add(company)
        occ_key = occ_name if occ_name else "(No Occupancy Classification / Standard Rate)"

        data.setdefault(company, {}).setdefault(product, {}).setdefault(occ_key, []).append(
            {
                "code": str(occ_code) if occ_code else "",
                "cat": str(category) if category else "",
                "recv": str(recv) if recv is not None else "",
                "pay": str(pay) if pay is not None else "",
            }
        )
        row_count += 1

    payload = {"companies": sorted(companies_set), "data": data}
    print(f"Read {row_count} data rows across {len(companies_set)} companies.")
    return json.dumps(payload, separators=(",", ":"))


def main():
    parser = argparse.ArgumentParser(description="Regenerate the SME Rate Lookup Tool's index.html")
    parser.add_argument("excel_file", help="Path to the current SME Compensation Matrix .xlsx file")
    parser.add_argument(
        "-o", "--output", default="index.html", help="Output HTML filename (default: index.html)"
    )
    args = parser.parse_args()

    xlsx_path = Path(args.excel_file)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    script_dir = Path(__file__).parent
    template_path = script_dir / TEMPLATE_FILENAME
    if not template_path.exists():
        sys.exit(
            f"Could not find {TEMPLATE_FILENAME} next to this script.\n"
            f"Expected it at: {template_path}"
        )

    print(f"Reading data from: {xlsx_path}")
    data_json = build_lookup_json(xlsx_path)

    if "</script>" in data_json:
        sys.exit(
            "Safety check failed: embedded data unexpectedly contains a "
            "'</script>' sequence. Please report this before proceeding."
        )

    template_html = template_path.read_text(encoding="utf-8")
    if PLACEHOLDER not in template_html:
        sys.exit(f"Placeholder {PLACEHOLDER} not found in {TEMPLATE_FILENAME}. Has it been edited?")

    final_html = template_html.replace(PLACEHOLDER, data_json)

    out_path = Path(args.output)
    out_path.write_text(final_html, encoding="utf-8")

    size_mb = out_path.stat().st_size / (1024 * 1024)
    print(f"Done. Wrote {out_path} ({size_mb:.2f} MB).")
    print("Upload this file to your host to replace the old version.")


if __name__ == "__main__":
    main()
